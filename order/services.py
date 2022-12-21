from datetime import datetime 

import openpyxl

from uuid import uuid4

from django.db.models import Model, QuerySet
from django.http import HttpRequest, HttpResponse
from django.utils.translation import gettext_lazy as _

from order.models import Order, OrderItem

from shop.models import Item


class OrderInfo:
	''' 주문정보 생성하는 class '''

	def __init__(self, request: HttpRequest) -> None:
		self.request = request

	def get_instance(self) -> Model:
		''' create order instance '''

		item_qs = Item.objects.filter(id__in=self.request.session['order_items'].keys())
		quantity_dict = { int(k): int(v) for k, v in self.request.session['order_items'].items() }

		order_items = list()
		for item in item_qs:
			quantity = quantity_dict[item.pk]
			order_item = OrderItem(quantity=quantity, item=item)
			order_items.append(order_item)

		user = self.request.user
		item = order_items[0]
		name = f'{item.item.name} 외 {len(order_items) - 1}건' \
			if len(order_items) > 1 else item.item.name

		instance = Order(
			name=name, 
			amount=sum(order_item.amount for order_item in order_items),
			buyer_name= f'{user.last_name}{user.first_name}',
			merchant_uid=uuid4,
			buyer_postcode=user.postal_code,
			buyer_addr=user.address,
			detail_addr=user.detail_address,
			buyer_email=user.email,
			buyer_tel=user.phone,
		)

		return order_items, instance


class OrderItemHandler:
	''' 주문상품을 handling하는 class '''

	def __init__(self, order: Model) -> None:
		self.order = order
		self.order_items = order.orderitem_order.all()

	def create_order_items(self, order_items: list[Model]) -> None:
		''' 주문상품을 bulk_create로 생성 '''

		for order_item in order_items:
			order_item.order = self.order
			order_item.status = self.order.status

		OrderItem.objects.bulk_create(order_items)

	def set_paid_status(self) -> None:
		''' 주문상품의 status를 order의 status(paid)로 변경 '''

		if not self.order.status == 'exchange_return':
			self.order_items.update(status=self.order.status)

	def set_cancel_status(self) -> None:
		''' order_item의 status를 주문취소로 변경 '''

		self.order_items.update(status='cancelled')

	def stock_minus_counter(self) -> None:
		''' item 수량을 order_item 수만큼 차감 '''

		if self.order.status == 'paid':
			for order_item in self.order_items:
				item = order_item.item
				item.stock = int(item.stock) - int(order_item.quantity)

				if item.stock == 0:
					item.is_public = False
				item.save()

	def stock_plus_counter(self) -> None:
		''' item 수량을 order_item 수만큼 가감 '''

		for order_item in self.order_items:
			item = order_item.item
			item.stock = int(item.stock) + int(order_item.quantity)

			if item.stock > 0:
				item.is_public = True
			item.save()

def export_excel(queryset: QuerySet) -> HttpResponse:
	''' 주문정보 엑셀로 다운로드 '''

	#TODO: 데이터 양이 많아지면 처리하는 시간이 많이 소요될것이라 celery나 asgi를 할용해서 비동기 처리가 필요할 것으로 생각됨 
	today = datetime.today().strftime('%Y%m%d')

	response = HttpResponse(content_type='application/ms-excel')
	response['Content-Disposition'] = f'attachment; filename="order_{today}.xlsx"'

	wb = openpyxl.Workbook()
	sheet = wb.active
	sheet.title = 'Order List'

	row_num = 1

	base_columns = {
		'id': _('id'),
		'user__username': _('구매자'),
		'merchant_uid': _('주문번호'),
		'imp_uid': _('아임포트_거래고유번호'),
		'name': _('상품명'),
		'amount': _('결제금액'),
		'delivery_amount': _('배송비'),
		'buyer_name': _('이름'),
		'buyer_postcode': _('우편번호'),
		'buyer_addr': _('주소'),
		'detail_addr': _('상세주소'),
		'buyer_email': _('이메일'),
		'buyer_tel': _('전화번호'),
		'exchange_return': _('교환/반품'),
		'delivery_number': _('등기번호'),
		'status': _('처리상태'),
		'pay_method': _('결제수단'),
		'vbank_num': _('가상계좌')
	}

	order_items_columns = {'order_items':  _('주문상품')}
	columns = dict(base_columns, **order_items_columns)

	for idx, col in enumerate(list(columns.keys()), start=1):
		sheet.cell(row=row_num, column=idx).value = str(columns[col])

	rows = queryset.values_list(*(list(base_columns.keys())))

	for row in rows:
		order_items = OrderItem.objects.filter(order__id=row[0]).values_list('item__name', 'quantity', 'status',)
		order_item = ','.join(map(str, order_items))
		row_num += 1

		for col_num in range(len(row)):
			sheet.cell(row=row_num, column=col_num+1).value = str(row[col_num])

		sheet.cell(row=row_num, column=len(row)+1).value = str(order_item)

	wb.save(response)
	return response

def import_excel(file_name):
	''' 등기번호 엑셀에서 불러오기(업데이트) '''

	wb = openpyxl.load_workbook(filename=file_name)
	sheet = wb.active
	max_row = sheet.max_row

	order_bulk_update_list = list()

	for i in range(2, max_row):
		order = Order.objects.get_or_none(pk=sheet.cell(row=i, column=1).value)

		if order is not None:
			value = sheet.cell(row=i, column=15).value if sheet.cell(row=i, column=15).value is not None else ''
			order.delivery_number = value

		order_bulk_update_list.append(order)

	Order.objects.bulk_update(order_bulk_update_list, ['delivery_number'])